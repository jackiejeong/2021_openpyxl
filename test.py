from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import sys, numpy as np

wb = load_workbook('D:\정성원\코딩\openpyxl\키워트\E2021_test4.xlsx')

# 시트 변수 지정
전체 = wb['전체']
Sheet1 = wb['Sheet1']
Sheet2 = wb['Sheet2']

# 판다스 데이터 변환 후 불러오기
df = pd.DataFrame(전체.values)
df.columns = df.iloc[0,:]
df = df.iloc[1:,:]

# 데이터 전처리
df_new = df[['국가코드', '발명의 명칭', '대표청구항', '요약', '독립항', '법적상태', '출원인', '출원인 국적',
             '출원번호', '출원일', '공개번호', '공개일', '등록번호', '등록일', '존속기간 만료일', '원출원번호',
             '원출원일자', '메인 IPC', '전체 IPC', '메인 CPC', '전체 CPC', '테마코드(JP)', 'F-Term(JP)', '우선권 국가',
             '우선권 번호', '우선권 주장일', '최우선일', '인용문헌수', '피인용문헌수', '패밀리 국가수', '패밀리 문헌수']]

# 출원년도 정리
출원년도 = []
for date in df_new['출원일']:
    year = date[0:4]
    출원년도.append(year)
    
# 출원연도 열 입력
df_new.insert(9, '출원년도', 출원년도)

# Sheet1 데이터 정리
출원년도counts = pd.DataFrame(df_new['출원년도'].value_counts())
Sheet1data = 출원년도counts.reset_index()
Sheet1data.columns = ['출원년도', '출원건수']
Sheet1data = Sheet1data.sort_values(by='출원년도', ascending = True)
Sheet1data['누적건수'] = np.cumsum(Sheet1data['출원건수'])
Sheet1data['연도'] = Sheet1data['출원년도'].str[2:]
Sheet1data = Sheet1data.reindex(columns=['출원년도', '연도', '출원건수', '누적건수'])

# Sheet1 데이터 쓰기
rows = dataframe_to_rows(Sheet1data, index=False)
for r_idx, row in enumerate(rows, 3):
    for c_idx, value in enumerate(row, 1):
        Sheet1.cell(row = r_idx, column = c_idx, value = value)
        
# Sheet1.insert_cols(2)
# for row, cellobj in enumerate(list(Sheet1.columns)[1]):
#     n = '=right(A%d,2)' % (row+1)
#     cellobj.value = n

# Sheet2 데이터 정리(한국/일본/미국/유럽)
Sheet2PT = pd.pivot_table(df_new, values = '출원번호', index = '출원년도', columns = '국가코드', aggfunc = 'count')
Sheet2PT = Sheet2PT.reindex(columns=['KR', 'JP', 'US', 'EP'])
# Sheet2PTpd = pd.DataFrame(Sheet2PT.values)


# Sheet2 데이터 쓰기
rows = dataframe_to_rows(Sheet2PT)
for r_idx, row in enumerate(rows, 3):
    for c_idx, value in enumerate(row, 1):
        Sheet2.cell(row = r_idx, column = c_idx, value = value)



# Sheet1PT = pd.pivot_table(df, values = '출원번호', index = '출원연도2', aggfunc = 'count')

        


wb.save('D:\정성원\코딩\openpyxl\키워트\E2021_test5.xlsx')