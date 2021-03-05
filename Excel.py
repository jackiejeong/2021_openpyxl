from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.chart import (LineChart, BarChart, PieChart, Reference, Series, reference, title)
from openpyxl.chart.label import DataLabelList
import tkinter.messagebox
import Data

wb = Workbook()

def ExcelGraph():
    A()
    B()
    C()
    D()
    E()
    # if Data.rb > 1:
    #     F()
    
def A():
    sheetA = wb.create_sheet('전체 출원동향', 0)
    A그래프data = Data.전체출원동향()

    for r in dataframe_to_rows(A그래프data, index=False, header=True):
        sheetA.append(r)

    sheetA.insert_cols(2)
    for row, cellobj in enumerate(list(sheetA.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    chartA1 = BarChart()
    dataA1 = Reference(sheetA, min_col = 3, min_row = 1, max_row = 21)
    catsA1 = Reference(sheetA, min_col = 2, min_row = 2, max_row = 21)
    chartA1.add_data(dataA1, titles_from_data = True)
    chartA1.set_categories(catsA1)
    chartA1.y_axis.majorGridlines = None

    chartA2 = LineChart()
    dataA2 = Reference(sheetA, min_col = 4, min_row = 1, max_row = 21)
    chartA2.add_data(dataA2, titles_from_data = True)
    chartA2.y_axis.majorGridlines = None
    chartA2.y_axis.axId = 2000

    # y축 위치 변경
    chartA2.y_axis.crosses = 'max'
    # 그래프 합치기
    chartA1 += chartA2
    chartA1.width = 15
    chartA1.height = 10
    chartA1.legend.position = 't'
    chartA1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True)) # 테두리 제거
    sheetA.add_chart(chartA1, 'F2')

    global savepath
    savepath = Data.Save()

def B():
    sheetB = wb.create_sheet('주요국가별 출원동향', 1)
    B그래프data = Data.주요국출원동향()

    for r in dataframe_to_rows(B그래프data, index=False, header=True):
        sheetB.append(r)

    sheetB.insert_cols(2)
    for row, cellobj in enumerate(list(sheetB.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    sheetB['B22'] = '합계'
    sheetB['C22'] = '=SUM(C2:C21)'
    sheetB['D22'] = '=SUM(D2:D21)'
    sheetB['E22'] = '=SUM(E2:E21)'
    sheetB['F22'] = '=SUM(F2:F21)'

    chartB1 = LineChart()
    dataB1 = Reference(sheetB, min_col = 3, min_row = 1, max_row = 21, max_col = 6)
    catsB1 = Reference(sheetB, min_col = 2, min_row = 2, max_row = 21)
    chartB1.add_data(dataB1, titles_from_data = True)
    chartB1.set_categories(catsB1)
    chartB1.y_axis.majorGridlines = None
    chartB1.width = 15
    chartB1.height = 10
    chartB1.legend.position = 't'
    chartB1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetB.add_chart(chartB1, 'H2')

    chartB2 = PieChart()
    dataB2 = Reference(sheetB, min_col = 3, min_row = 22, max_col = 6)
    labelsB2 = Reference(sheetB, min_col = 3, min_row = 1, max_col = 6)
    chartB2.add_data(dataB2, from_rows = 22, titles_from_data = False)
    chartB2.set_categories(labelsB2)
    chartB2.width = 5
    chartB2.height =5
    chartB2.legend = None
    chartB2.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True, solidFill = None, gradFill = None))
    chartB2.dLbls = DataLabelList()
    chartB2.dLbls.showPercent = True

    sheetB.add_chart(chartB2, 'M1')

def C():
    # 데이터 정리
    sheetC = wb.create_sheet('주요국 내 상위다출원국가', 2)
    C상위다출원국가KR = Data.상위다출원국가('KR')
    C상위다출원국가JP = Data.상위다출원국가('JP')
    C상위다출원국가US = Data.상위다출원국가('US')
    C상위다출원국가EP = Data.상위다출원국가('EP')

    for r in dataframe_to_rows(C상위다출원국가KR, index=False, header=True):
        sheetC.append(r)

    sheetC['A22'] = ''
    sheetC['A23'] = ''

    for r in dataframe_to_rows(C상위다출원국가JP, index=False, header=True):
        sheetC.append(r)

    sheetC['A45'] = ''
    sheetC['A46'] = ''

    for r in dataframe_to_rows(C상위다출원국가US, index=False, header=True):
        sheetC.append(r)

    sheetC['A68'] = ''
    sheetC['A69'] = ''

    for r in dataframe_to_rows(C상위다출원국가EP, index=False, header=True):
        sheetC.append(r)

    sheetC['A91'] = ''
    sheetC['A92'] = ''

    sheetC.insert_cols(2)
    for row, cellobj in enumerate(list(sheetC.columns)[1]):
        n = '=right(A%d,2)' % (row+1)
        cellobj.value = n

    # 엑셀 함수
    sheetC['B22'] = '합계'
    sheetC['C22'] = '=SUM(C2:C21)'
    sheetC['D22'] = '=SUM(D2:D21)'
    sheetC['E22'] = '=SUM(E2:E21)'
    sheetC['F22'] = '=SUM(F2:F21)'
    sheetC['A23'] = ' '

    sheetC['B45'] = '합계'
    sheetC['C45'] = '=SUM(C25:C44)'
    sheetC['D45'] = '=SUM(D25:D44)'
    sheetC['E45'] = '=SUM(E25:E44)'
    sheetC['F45'] = '=SUM(F25:F44)'
    sheetC['A46'] = ' '

    sheetC['B68'] = '합계'
    sheetC['C68'] = '=SUM(C48:C67)'
    sheetC['D68'] = '=SUM(D48:D67)'
    sheetC['E68'] = '=SUM(E48:E67)'
    sheetC['F68'] = '=SUM(F48:F67)'
    sheetC['A69'] = ' '

    sheetC['B91'] = '합계'
    sheetC['C91'] = '=SUM(C71:C90)'
    sheetC['D91'] = '=SUM(D71:D90)'
    sheetC['E91'] = '=SUM(E71:E90)'
    sheetC['F91'] = '=SUM(F71:F90)'
    sheetC['A92'] = ' '

    # 그래프 그리기
    chartC11 = LineChart()
    dataC11 = Reference(sheetC, min_col = 3, min_row = 1, max_col = 6, max_row = 21)
    catsC11 = Reference(sheetC, min_col = 2, min_row = 2, max_row = 21)
    chartC11.add_data(dataC11, titles_from_data = True)
    chartC11.set_categories(catsC11)
    chartC11.y_axis.majorGridlines = None
    chartC11.width = 15
    chartC11.height = 10
    chartC11.legend.position = 't'
    chartC11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC11, 'H2')

    chartC12 = PieChart()
    dataC12 = Reference(sheetC, min_col = 3, min_row = 22, max_col = 6)
    labelC12 = Reference(sheetC, min_col = 3, min_row = 1, max_col = 6)
    chartC12.add_data(dataC12, from_rows = 22, titles_from_data = False)
    chartC12.set_categories(labelC12)
    chartC12.width = 5
    chartC12.height = 5
    chartC12.legend = None
    chartC12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC12.dLbls = DataLabelList()
    chartC12.dLbls.showPercent = True
    sheetC.add_chart(chartC12, 'M1')

    chartC21 = LineChart()
    dataC21 = Reference(sheetC, min_col = 3, min_row = 24, max_col = 6, max_row = 44)
    catsC21 = Reference(sheetC, min_col = 2, min_row = 25, max_row = 44)
    chartC21.add_data(dataC21, titles_from_data = True)
    chartC21.set_categories(catsC21)
    chartC21.y_axis.majorGridlines = None
    chartC21.width = 15
    chartC21.height = 10
    chartC21.legend.position = 't'
    chartC21.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC21, 'H25')

    chartC22 = PieChart()
    dataC22 = Reference(sheetC, min_col = 3, min_row = 45, max_col = 6)
    labelC22 = Reference(sheetC, min_col = 3, min_row = 24, max_col = 6)
    chartC22.add_data(dataC22, from_rows = 45, titles_from_data = False)
    chartC22.set_categories(labelC22)
    chartC22.width = 5
    chartC22.height = 5
    chartC22.legend = None
    chartC22.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC22.dLbls = DataLabelList()
    chartC22.dLbls.showPercent = True
    sheetC.add_chart(chartC22, 'M24')

    chartC31 = LineChart()
    dataC31 = Reference(sheetC, min_col = 3, min_row = 47, max_col = 6, max_row = 67)
    catsC31 = Reference(sheetC, min_col = 2, min_row = 48, max_row = 67)
    chartC31.add_data(dataC31, titles_from_data = True)
    chartC31.set_categories(catsC31)
    chartC31.y_axis.majorGridlines = None
    chartC31.width = 15
    chartC31.height = 10
    chartC31.legend.position = 't'
    chartC31.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC31, 'H48')

    chartC32 = PieChart()
    dataC32 = Reference(sheetC, min_col = 3, min_row = 68, max_col = 6)
    labelC32 = Reference(sheetC, min_col = 3, min_row = 47, max_col = 6)
    chartC32.add_data(dataC32, from_rows = 68, titles_from_data = False)
    chartC32.set_categories(labelC32)
    chartC32.width = 5
    chartC32.height = 5
    chartC32.legend = None
    chartC32.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC32.dLbls = DataLabelList()
    chartC32.dLbls.showPercent = True
    sheetC.add_chart(chartC32, 'M47')

    chartC41 = LineChart()
    dataC41 = Reference(sheetC, min_col = 3, min_row = 70, max_col = 6, max_row = 90)
    catsC41 = Reference(sheetC, min_col = 2, min_row = 71, max_row = 90)
    chartC41.add_data(dataC41, titles_from_data = True)
    chartC41.set_categories(catsC41)
    chartC41.y_axis.majorGridlines = None
    chartC41.width = 15
    chartC41.height = 10
    chartC41.legend.position = 't'
    chartC41.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    sheetC.add_chart(chartC41, 'H71')

    chartC42 = PieChart()
    dataC42 = Reference(sheetC, min_col = 3, min_row = 91, max_col = 6)
    labelC42 = Reference(sheetC, min_col = 3, min_row = 70, max_col = 6)
    chartC42.add_data(dataC42, from_rows = 91, titles_from_data = False)
    chartC42.set_categories(labelC42)
    chartC42.width = 5
    chartC42.height = 5
    chartC42.legend = None
    chartC42.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartC42.dLbls = DataLabelList()
    chartC42.dLbls.showPercent = True
    sheetC.add_chart(chartC42, 'M70')

    # if Data.rb == 1:
    #     wb.save('{}.xlsx'.format(savepath))
    #     tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류X)')

def D():
    sheetD = wb.create_sheet('주요국 내외국인 출원점유율', 3)
    D그래프data1 = Data.내외국인점유율()
    for r in dataframe_to_rows(D그래프data1, index=False, header=True):
        sheetD.append(r)
    sheetD['A4'] = '합계'
    sheetD['B4'] = '=SUM(B2:B3)'
    sheetD['C4'] = '=SUM(C2:C3)'
    sheetD['D4'] = '=SUM(D2:D3)'
    sheetD['E4'] = '=SUM(E2:E3)'
    sheetD['A5'] = ' '

    D그래프data2 = Data.외국인점유율()
    for r in dataframe_to_rows(D그래프data2, index=False, header=True):
        sheetD.append(r)

    # 전체 그래프
    chartD1 = PieChart()
    dataD1 = Reference(sheetD, min_col = 2, min_row = 4, max_col = 5)
    labelD1 = Reference(sheetD, min_col = 2, min_row = 1, max_col = 5)
    chartD1.add_data(dataD1, from_rows = 4, titles_from_data = False)
    chartD1.set_categories(labelD1)
    chartD1.width = 8
    chartD1.height = 8
    chartD1.legend = None
    chartD1.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD1.dLbls = DataLabelList()
    chartD1.dLbls.showPercent = True
    chartD1.dLbls.showCatName = True
    sheetD.add_chart(chartD1, 'P2')

    # KR 그래프
    chartD2 = PieChart()
    dataD2 = Reference(sheetD, min_col = 2, min_row = 2, max_row = 3)
    labelD2 = Reference(sheetD, min_col = 1, min_row = 2, max_row = 3)
    chartD2.add_data(dataD2, titles_from_data = False)
    chartD2.set_categories(labelD2)
    chartD2.width = 6.5
    chartD2.height = 6.5
    chartD2.legend = None
    chartD2.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD2.dLbls = DataLabelList()
    chartD2.dLbls.showPercent = True
    chartD2.dLbls.showCatName = True
    sheetD.add_chart(chartD2, 'K18')

    # JP 그래프
    chartD3 = PieChart()
    dataD3 = Reference(sheetD, min_col = 3, min_row = 2, max_row = 3)
    labelD3 = Reference(sheetD, min_col = 1, min_row = 2, max_row = 3)
    chartD3.add_data(dataD3, titles_from_data = False)
    chartD3.set_categories(labelD3)
    chartD3.width = 6.5
    chartD3.height = 6.5
    chartD3.legend = None
    chartD3.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD3.dLbls = DataLabelList()
    chartD3.dLbls.showPercent = True
    chartD3.dLbls.showCatName = True
    sheetD.add_chart(chartD3, 'O18')

    # US 그래프
    chartD4 = PieChart()
    dataD4 = Reference(sheetD, min_col = 4, min_row = 2, max_row = 3)
    labelD4 = Reference(sheetD, min_col = 1, min_row = 2, max_row = 3)
    chartD4.add_data(dataD4, titles_from_data = False)
    chartD4.set_categories(labelD4)
    chartD4.width = 6.5
    chartD4.height = 6.5
    chartD4.legend = None
    chartD4.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD4.dLbls = DataLabelList()
    chartD4.dLbls.showPercent = True
    chartD4.dLbls.showCatName = True
    sheetD.add_chart(chartD4, 'S18')

    # EP 그래프
    chartD5 = PieChart()
    dataD5 = Reference(sheetD, min_col = 5, min_row = 2, max_row = 3)
    labelD5 = Reference(sheetD, min_col = 1, min_row = 2, max_row = 3)
    chartD5.add_data(dataD5, titles_from_data = False)
    chartD5.set_categories(labelD5)
    chartD5.width = 6.5
    chartD5.height = 6.5
    chartD5.legend = None
    chartD5.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD5.dLbls = DataLabelList()
    chartD5.dLbls.showPercent = True
    chartD5.dLbls.showCatName = True
    sheetD.add_chart(chartD5, 'W18')

    # KR 외국인 그래프(상위 4개국)
    chartD6 = PieChart()
    dataD6 = Reference(sheetD, min_col = 2, min_row = 7, max_row = 10)
    labelD6 = Reference(sheetD, min_col = 1, min_row = 7, max_row = 10)
    chartD6.add_data(dataD6, titles_from_data = False)
    chartD6.set_categories(labelD6)
    chartD6.width = 6.5
    chartD6.height = 6.5
    chartD6.legend = None
    chartD6.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD6.dLbls = DataLabelList()
    chartD6.dLbls.showPercent = True
    chartD6.dLbls.showCatName = True
    sheetD.add_chart(chartD6, 'K31')

    # JP 외국인 그래프(상위 4개국)
    chartD7 = PieChart()
    dataD7 = Reference(sheetD, min_col = 4, min_row = 7, max_row = 10)
    labelD7 = Reference(sheetD, min_col = 3, min_row = 7, max_row = 10)
    chartD7.add_data(dataD7, titles_from_data = False)
    chartD7.set_categories(labelD7)
    chartD7.width = 6.5
    chartD7.height = 6.5
    chartD7.legend = None
    chartD7.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD7.dLbls = DataLabelList()
    chartD7.dLbls.showPercent = True
    chartD7.dLbls.showCatName = True
    sheetD.add_chart(chartD7, 'O31')

    # US 외국인 그래프(상위 4개국)
    chartD8 = PieChart()
    dataD8 = Reference(sheetD, min_col = 6, min_row = 7, max_row = 10)
    labelD8 = Reference(sheetD, min_col = 5, min_row = 7, max_row = 10)
    chartD8.add_data(dataD8, titles_from_data = False)
    chartD8.set_categories(labelD8)
    chartD8.width = 6.5
    chartD8.height = 6.5
    chartD8.legend = None
    chartD8.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD8.dLbls = DataLabelList()
    chartD8.dLbls.showPercent = True
    chartD8.dLbls.showCatName = True
    sheetD.add_chart(chartD8, 'S31')

    # EP 외국인 그래프(상위 4개국)
    chartD9 = PieChart()
    dataD9 = Reference(sheetD, min_col = 8, min_row = 7, max_row = 10)
    labelD9 = Reference(sheetD, min_col = 7, min_row = 7, max_row = 10)
    chartD9.add_data(dataD9, titles_from_data = False)
    chartD9.set_categories(labelD9)
    chartD9.width = 6.5
    chartD9.height = 6.5
    chartD9.legend = None
    chartD9.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
    chartD9.dLbls = DataLabelList()
    chartD9.dLbls.showPercent = True
    chartD9.dLbls.showCatName = True
    sheetD.add_chart(chartD9, 'W31')

    # wb.save('{}.xlsx'.format(savepath))
    # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료')

def E():
    E그래프data = Data.기술성장()
    sheetE = wb.create_sheet('기술성장도', 4)
    D그래프data1 = Data.내외국인점유율()
    for r in dataframe_to_rows(E그래프data, index=False, header=True):
        sheetE.append(r)



    wb.save('{}.xlsx'.format(savepath))
    tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료')



# def F():
#     sheetF = wb.create_sheet('기술분류별 출원동향', ?)
#     F그래프data = Data.기술분류()
#     for r in dataframe_to_rows(F그래프data, index=False, header=True):
#         sheetF.append(r)

#     sheetF.insert_cols(2)
#     for row, cellobj in enumerate(list(sheetF.columns)[1]):
#         n = '=right(A%d,2)' % (row+1)
#         cellobj.value = n

#     sheetF['B22'] = '합계'
#     sheetF['C22'] = '=SUM(C2:C21)'
#     sheetF['D22'] = '=SUM(D2:D21)'
#     sheetF['E22'] = '=SUM(E2:E21)'
#     sheetF['F22'] = '=SUM(F2:F21)'

#     if Data.rb == 2:
#         chartF11 = LineChart()
#         dataF11 = Reference(sheetF, min_col = 3, min_row = 1, max_col = 4, max_row = 21)
#         catsF11 = Reference(sheetF, min_col = 2, min_row = 2, max_row = 21)
#         chartF11.add_data(dataF11, titles_from_data = True)
#         chartF11.set_categories(catsF11)
#         chartF11.y_axis.majorGridlines = None
#         chartF11.width = 15
#         chartF11.height = 10
#         chartF11.legend.position = 't'
#         chartF11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
#         sheetF.add_chart(chartF11, 'H2')
        
#         chartF12 = PieChart()
#         dataF12 = Reference(sheetF, min_col = 3, min_row = 22, max_col = 4)
#         labelF12 = Reference(sheetF, min_col = 3, min_row = 1, max_col = 4)
#         chartF12.add_data(dataF12, from_rows = 22, titles_from_data = False)
#         chartF12.set_categories(labelF12)
#         chartF12.width = 5
#         chartF12.height = 5
#         chartF12.legend = None
#         chartF12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
#         chartF12.dLbls = DataLabelList()
#         chartF12.dLbls.showPercent = True
#         sheetF.add_chart(chartF12, 'M1')

#         # wb.save('{}.xlsx'.format(savepath))
#         # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류 2개)')

#     elif Data.rb == 3:
#         chartF11 = LineChart()
#         dataF11 = Reference(sheetF, min_col = 3, min_row = 1, max_col = 5, max_row = 21)
#         catsF11 = Reference(sheetF, min_col = 2, min_row = 2, max_row = 21)
#         chartF11.add_data(dataF11, titles_from_data = True)
#         chartF11.set_categories(catsF11)
#         chartF11.y_axis.majorGridlines = None
#         chartF11.width = 15
#         chartF11.height = 10
#         chartF11.legend.position = 't'
#         chartF11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
#         sheetF.add_chart(chartF11, 'H2')
        
#         chartF12 = PieChart()
#         dataF12 = Reference(sheetF, min_col = 3, min_row = 22, max_col = 5)
#         labelF12 = Reference(sheetF, min_col = 3, min_row = 1, max_col = 5)
#         chartF12.add_data(dataF12, from_rows = 22, titles_from_data = False)
#         chartF12.set_categories(labelF12)
#         chartF12.width = 5
#         chartF12.height = 5
#         chartF12.legend = None
#         chartF12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
#         chartF12.dLbls = DataLabelList()
#         chartF12.dLbls.showPercent = True
#         sheetF.add_chart(chartF12, 'M1')

#         # wb.save('{}.xlsx'.format(savepath))
#         # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류 3개)')

#     elif Data.rb == 4:
#         chartF11 = LineChart()
#         dataF11 = Reference(sheetF, min_col = 3, min_row = 1, max_col = 6, max_row = 21)
#         catsF11 = Reference(sheetF, min_col = 2, min_row = 2, max_row = 21)
#         chartF11.add_data(dataF11, titles_from_data = True)
#         chartF11.set_categories(catsF11)
#         chartF11.y_axis.majorGridlines = None
#         chartF11.width = 15
#         chartF11.height = 10
#         chartF11.legend.position = 't'
#         chartF11.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
#         sheetF.add_chart(chartF11, 'H2')
        
#         chartF12 = PieChart()
#         dataF12 = Reference(sheetF, min_col = 3, min_row = 22, max_col = 6)
#         labelF12 = Reference(sheetF, min_col = 3, min_row = 1, max_col = 6)
#         chartF12.add_data(dataF12, from_rows = 22, titles_from_data = False)
#         chartF12.set_categories(labelF12)
#         chartF12.width = 5
#         chartF12.height = 5
#         chartF12.legend = None
#         chartF12.graphical_properties = GraphicalProperties(ln=LineProperties(noFill=True))
#         chartF12.dLbls = DataLabelList()
#         chartF12.dLbls.showPercent = True
#         sheetF.add_chart(chartF12, 'M1')

#         # wb.save('{}.xlsx'.format(savepath))
#         # tkinter.messagebox.showinfo('messagebox', '그래프 생성 완료(기술분류 4개)')
